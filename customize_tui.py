# -*- coding: utf-8 -*-
"""
NCSS-Nav 一键换校工具 v2.0
===========================
TUI交互界面 · Excel模板导入 · 一键生成.exe

功能：
  1. 交互式问答配置
  2. Excel模板导入
  3. JSON配置文件导入
  4. 预览配置预览
  5. 一键执行适配
"""
import os
import sys
import json
import re
import shutil
from pathlib import Path
from datetime import datetime

try:
    from rich.console import Console
    from rich.table import Table
    from rich.panel import Panel
    from rich.prompt import Prompt, Confirm
    from rich.layout import Layout
    from rich.text import Text
    from rich.columns import Columns
    from rich.box import ROUNDED, HEAVY
    from rich.style import Style
    HAS_RICH = True
except ImportError:
    HAS_RICH = False
    print("提示：安装 rich 库可获得更好的界面体验：pip install rich")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── 路径配置 ────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
SRC_DIR = BASE_DIR / 'src'
CONFIG_FILE = BASE_DIR / 'src' / 'config' / 'site.js'
APPS_FILE = BASE_DIR / 'src' / 'data' / 'apps.js'
TEMPLATES_DIR = BASE_DIR / 'templates'

# ─── 控制台 ──────────────────────────────────────────────
console = Console() if HAS_RICH else None

def print_rich(*args, **kwargs):
    if console:
        console.print(*args, **kwargs)
    else:
        print(*args)

# ─── 颜色主题预设 ────────────────────────────────────────────
THEMES = {
    '1': {'name': '闽都红（默认）', 'primary': '#c62828', 'accent': '#ff5722', 'bg': '#fef2f2'},
    '2': {'name': '北大红', 'primary': '#8b0000', 'accent': '#cc0000', 'bg': '#fff5f5'},
    '3': {'name': '清华紫', 'primary': '#6b21a8', 'accent': '#9333ea', 'bg': '#faf5ff'},
    '4': {'name': '复旦蓝', 'primary': '#1e40af', 'accent': '#3b82f6', 'bg': '#eff6ff'},
    '5': {'name': '浙大绿', 'primary': '#166534', 'accent': '#22c55e', 'bg': '#f0fdf4'},
    '6': {'name': '交大橙', 'primary': '#c2410c', 'accent': '#f97316', 'bg': '#fff7ed'},
    '7': {'name': '武大樱', 'primary': '#be185d', 'accent': '#ec4899', 'bg': '#fdf2f8'},
    '8': {'name': '华科灰', 'primary': '#374151', 'accent': '#6b7280', 'bg': '#f9fafb'},
    '9': {'name': '自定义', 'primary': '', 'accent': '', 'bg': ''},
}

# ─── Excel列名映射 ────────────────────────────────────────────
EXCEL_COLUMNS = {
    '学校全称': 'university',
    '学校简称': 'shortName', 
    '品牌名': 'brand',
    '校训': 'motto',
    '教务处网址': 'jwUrl',
    '贴吧URL': 'tiebaUrl',
    'GitHub仓库': 'github',
    'Pages地址': 'pagesUrl',
    '校区数量': 'campuses',
    '学院数量': 'colleges',
    '专业数量': 'majors',
    '主题编号': 'theme_id',
    '主色调': 'primary',
    '强调色': 'accent',
    '背景色': 'bg',
}

def show_banner():
    """显示欢迎横幅"""
    if console:
        banner = Panel(
            Text("NCSS-Nav 一键换校工具 v2.0", style="bold cyan", justify="center"),
            subtitle="模块化校园导航 · 5分钟适配新高校",
            box=ROUNDED,
            border_style="bright_blue"
        )
        console.print(banner)
    else:
        print("=" * 60)
        print("  NCSS-Nav 一键换校工具 v2.0")
        print("  模块化校园导航 · 5分钟适配新高校")
        print("=" * 60)

def show_menu():
    """显示主菜单"""
    if console:
        table = Table(title="请选择配置方式", box=ROUNDED, border_style="cyan")
        table.add_column("编号", style="bold yellow", width=6)
        table.add_column("方式", style="bold white")
        table.add_column("说明", style="dim")
        table.add_row("1", "交互式配置", "通过问答逐步填写配置")
        table.add_row("2", "Excel模板导入", "下载模板Excel，填写后导入")
        table.add_row("3", "JSON配置导入", "从已有的JSON配置文件导入")
        table.add_row("4", "退出", "退出程序")
        console.print(table)
    else:
        print("\n请选择配置方式：")
        print("  1. 交互式配置")
        print("  2. Excel模板导入")
        print("  3. JSON配置导入")
        print("  4. 退出")

def input_with_default(prompt, default=''):
    """带默认值的输入"""
    if console:
        val = Prompt.ask(f"[cyan]{prompt}[/cyan]", default=default)
    else:
        val = input(f'{prompt} [{default}]: ').strip()
        if not val:
            val = default
    return val

def select_theme():
    """选择主题"""
    if console:
        table = Table(title="选择配色主题", box=ROUNDED, border_style="green")
        table.add_column("编号", style="bold yellow", width=6)
        table.add_column("主题名", style="bold white")
        table.add_column("主色调", style="bold")
        for k, v in THEMES.items():
            color = v['primary'] if v['primary'] else '默认'
            table.add_row(k, v['name'], f"[{color}]███[/{color}]" if v['primary'] else color)
        console.print(table)
    else:
        print('\n请选择配色主题：')
        for k, v in THEMES.items():
            print(f'  {k}. {v["name"]}')
    
    choice = input_with_default('输入编号 (1-9)', '1')
    if choice not in THEMES:
        choice = '1'
    theme = THEMES[choice]
    
    if choice == '9':
        theme['primary'] = input_with_default('主色调（十六进制）', '#c62828')
        theme['accent'] = input_with_default('强调色（十六进制）', '#ff5722')
        theme['bg'] = input_with_default('背景色（十六进制）', '#fef2f2')
    return theme

def gather_info_interactive():
    """交互式收集配置信息"""
    info = {}
    
    if console:
        console.print("\n[bold cyan]📝 基本信息[/bold cyan]")
    else:
        print('\n📝 基本信息：')
    
    info['university'] = input_with_default('学校全称', '福建师范大学')
    info['shortName'] = input_with_default('学校简称', '福star')
    info['brand'] = input_with_default('品牌名（英文/缩写）', 'FJNU-Nav')
    info['motto'] = input_with_default('校训', '知明行笃 · 立诚致广')
    
    if console:
        console.print("\n[bold cyan]🌐 网络信息[/bold cyan]")
    else:
        print('\n🌐 网络信息：')
    
    info['jwUrl'] = input_with_default('教务处网址', 'https://jwc.fjnu.edu.cn')
    info['tiebaUrl'] = input_with_default('贴吧URL', 'https://tieba.baidu.com/f?kw=%E7%A6%8F%E5%BB%BA%E5%B8%88%E8%8C%83%E5%A4%A7%E5%AD%A6')
    info['github'] = input_with_default('GitHub仓库地址', 'https://github.com/icyteacn/FJNU-Nav')
    info['pagesUrl'] = input_with_default('GitHub Pages地址', 'https://icyteacn.github.io/FJNU-Nav/')
    
    if console:
        console.print("\n[bold cyan]🏛️ 学校数据[/bold cyan]")
    else:
        print('\n🏛️ 学校数据：')
    
    info['campuses'] = input_with_default('校区数量', '2')
    info['colleges'] = input_with_default('学院数量', '28')
    info['majors'] = input_with_default('专业数量', '84')
    
    if console:
        console.print("\n[bold cyan]🎨 配色[/bold cyan]")
    else:
        print('\n🎨 配色：')
    
    info['theme'] = select_theme()
    
    return info

def generate_excel_template():
    """生成Excel模板"""
    if not HAS_OPENPYXL:
        print_rich("[red]错误：需要安装 openpyxl 库[/red]")
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "换校配置"
    
    # 设置标题样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["配置项", "配置值", "说明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 数据行
    data = [
        ("学校全称", "福建师范大学", "完整学校名称"),
        ("学校简称", "福star", "用于显示的简称"),
        ("品牌名", "FJNU-Nav", "英文/缩写品牌名"),
        ("校训", "知明行笃 · 立诚致广", "学校校训"),
        ("教务处网址", "https://jwc.fjnu.edu.cn", "教务处官网"),
        ("贴吧URL", "https://tieba.baidu.com/f?kw=福建师范大学", "学校贴吧"),
        ("GitHub仓库", "https://github.com/icyteacn/FJNU-Nav", "GitHub仓库地址"),
        ("Pages地址", "https://icyteacn.github.io/FJNU-Nav/", "GitHub Pages地址"),
        ("校区数量", "2", "数字"),
        ("学院数量", "28", "数字"),
        ("专业数量", "84", "数字"),
        ("主题编号", "1", "1-8预设，9自定义"),
        ("主色调", "#c62828", "仅主题9有效"),
        ("强调色", "#ff5722", "仅主题9有效"),
        ("背景色", "#fef2f2", "仅主题9有效"),
    ]
    
    for row, (key, value, desc) in enumerate(data, 2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=desc)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    
    # 保存
    template_path = BASE_DIR / "templates" / "换校模板.xlsx"
    template_path.parent.mkdir(exist_ok=True)
    wb.save(template_path)
    
    return template_path

def import_from_excel(excel_path):
    """从Excel导入配置"""
    if not HAS_OPENPYXL:
        print_rich("[red]错误：需要安装 openpyxl 库[/red]")
        return None
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    info = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).strip()
            value = str(row[1]).strip()
            if key in EXCEL_COLUMNS:
                info[EXCEL_COLUMNS[key]] = value
    
    # 处理主题
    theme_id = info.pop('theme_id', '1')
    if theme_id not in THEMES:
        theme_id = '1'
    theme = THEMES[theme_id].copy()
    
    # 自定义主题
    if theme_id == '9':
        theme['primary'] = info.pop('primary', '#c62828')
        theme['accent'] = info.pop('accent', '#ff5722')
        theme['bg'] = info.pop('bg', '#fef2f2')
    else:
        info.pop('primary', None)
        info.pop('accent', None)
        info.pop('bg', None)
    
    info['theme'] = theme
    return info

def import_from_json(json_path):
    """从JSON导入配置"""
    with open(json_path, 'r', encoding='utf-8') as f:
        info = json.load(f)
    if 'theme' not in info:
        info['theme'] = THEMES['1']
    return info

def preview_config(info):
    """预览配置"""
    if console:
        table = Table(title="配置预览", box=ROUNDED, border_style="yellow")
        table.add_column("配置项", style="bold cyan", width=15)
        table.add_column("配置值", style="white")
        
        table.add_row("学校全称", info.get('university', '-'))
        table.add_row("学校简称", info.get('shortName', '-'))
        table.add_row("品牌名", info.get('brand', '-'))
        table.add_row("校训", info.get('motto', '-'))
        table.add_row("教务处网址", info.get('jwUrl', '-'))
        table.add_row("校区数量", info.get('campuses', '-'))
        table.add_row("学院数量", info.get('colleges', '-'))
        table.add_row("专业数量", info.get('majors', '-'))
        
        theme = info.get('theme', {})
        table.add_row("主题", theme.get('name', '-'))
        table.add_row("主色调", theme.get('primary', '-'))
        
        console.print(table)
    else:
        print("\n配置预览：")
        for key, value in info.items():
            if key != 'theme':
                print(f"  {key}: {value}")
        print(f"  主题: {info.get('theme', {}).get('name', '-')}")

def update_site_config(info):
    """更新site.js"""
    content = CONFIG_FILE.read_text(encoding='utf-8')
    
    replacements = {
        r"brand:\s*'[^']*'": f"brand: '{info['brand']}'",
        r"name:\s*'[^']*'": f"name: '{info['shortName']} 校园导航'",
        r"tagline:\s*'[^']*'": f"tagline: '{info['university']} · 校园服务聚合入口'",
        r"motto:\s*'[^']*'": f"motto: '{info['motto']}'",
        r"copy:\s*'[^']*'": f"copy: '{info['shortName']} 校园导航 · 非官方校园服务聚合演示站，数据仅供学习交流'",
        r"jwUrl:\s*'[^']*'": f"jwUrl: '{info['jwUrl']}'",
    }
    
    for pattern, replacement in replacements.items():
        content = re.sub(pattern, replacement, content, count=1)
    
    CONFIG_FILE.write_text(content, encoding='utf-8')
    print_rich("[green]✅ site.js 已更新[/green]")

def update_apps_stats(info):
    """更新apps.js"""
    content = APPS_FILE.read_text(encoding='utf-8')
    content = re.sub(r"campuses:\s*\d+", f"campuses: {info['campuses']}", content)
    content = re.sub(r"colleges:\s*\d+", f"colleges: {info['colleges']}", content)
    content = re.sub(r"majors:\s*\d+", f"majors: {info['majors']}", content)
    APPS_FILE.write_text(content, encoding='utf-8')
    print_rich("[green]✅ apps.js 学校数据已更新[/green]")

def update_theme(info):
    """更新主题样式"""
    theme = info['theme']
    css_file = SRC_DIR / 'styles.css'
    if css_file.exists():
        content = css_file.read_text(encoding='utf-8')
        content = re.sub(r'--primary:\s*[^;]+;', f'--primary: {theme["primary"]};', content)
        content = re.sub(r'--accent:\s*[^;]+;', f'--accent: {theme["accent"]};', content)
        css_file.write_text(content, encoding='utf-8')
        print_rich("[green]✅ 样式主题已更新[/green]")

def update_tieba_url(info):
    """更新贴吧链接"""
    tieba_file = SRC_DIR / 'views' / 'TiebaSentiment.vue'
    if tieba_file.exists():
        content = tieba_file.read_text(encoding='utf-8')
        content = re.sub(
            r"https://tieba\.baidu\.com/f\?kw=[^'\"]+",
            info['tiebaUrl'],
            content
        )
        tieba_file.write_text(content, encoding='utf-8')
        print_rich("[green]✅ 贴吧链接已更新[/green]")

def update_sentiment_data(info):
    """更新本站舆情数据"""
    sentiment_file = SRC_DIR / 'data' / 'siteSentiment.js'
    if sentiment_file.exists():
        content = sentiment_file.read_text(encoding='utf-8')
        content = re.sub(r"FJNU-Nav", info['brand'], content)
        sentiment_file.write_text(content, encoding='utf-8')
        print_rich("[green]✅ 本站舆情品牌名已更新[/green]")

def update_industry_data(info):
    """更新产业价值数据"""
    industry_file = SRC_DIR / 'data' / 'industryValue.js'
    if industry_file.exists():
        content = industry_file.read_text(encoding='utf-8')
        content = re.sub(r"福建师范大学", info['university'], content)
        industry_file.write_text(content, encoding='utf-8')
        print_rich("[green]✅ 产业价值学校名已更新[/green]")

def execute_adaptation(info):
    """执行适配"""
    print_rich("\n[bold cyan]🔧 开始适配...[/bold cyan]")
    
    update_site_config(info)
    update_apps_stats(info)
    update_theme(info)
    update_tieba_url(info)
    update_sentiment_data(info)
    update_industry_data(info)
    
    # 保存配置文件
    config_path = TEMPLATES_DIR / f"{info['brand']}.json"
    TEMPLATES_DIR.mkdir(exist_ok=True)
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(info, f, ensure_ascii=False, indent=2)
    print_rich(f"[green]✅ 配置文件已保存: {config_path.name}[/green]")

def show_completion(info):
    """显示完成信息"""
    if console:
        panel = Panel(
            f"[bold green]适配完成！[/bold green]\n\n"
            f"学校：[cyan]{info['university']}[/cyan]\n"
            f"品牌：[cyan]{info['brand']}[/cyan]\n"
            f"主色：[{info['theme']['primary']}]███[/{info['theme']['primary']}] {info['theme']['primary']}\n\n"
            f"[bold]下一步：[/bold]\n"
            f"  1. npm install\n"
            f"  2. npm run dev\n"
            f"  3. 打开浏览器查看效果\n"
            f"  4. 根据需要修改 src/data/ 下的静态数据",
            title="🎉 适配完成",
            border_style="green"
        )
        console.print(panel)
    else:
        print('\n' + '=' * 60)
        print(f'  ✅ 适配完成！')
        print(f'  学校：{info["university"]}')
        print(f'  品牌：{info["brand"]}')
        print(f'  主色：{info["theme"]["primary"]}')
        print('=' * 60)
        print('\n下一步：')
        print('  1. npm install')
        print('  2. npm run dev')
        print('  3. 打开浏览器查看效果')

def main():
    """主函数"""
    show_banner()
    
    # 命令行参数处理
    if len(sys.argv) > 1:
        if sys.argv[1] == '--excel' and len(sys.argv) > 2:
            # Excel导入模式
            info = import_from_excel(sys.argv[2])
            if info:
                preview_config(info)
                if Confirm.ask("确认执行适配？"):
                    execute_adaptation(info)
                    show_completion(info)
            return
        elif sys.argv[1] == '--config' and len(sys.argv) > 2:
            # JSON导入模式
            info = import_from_json(sys.argv[2])
            preview_config(info)
            if Confirm.ask("确认执行适配？"):
                execute_adaptation(info)
                show_completion(info)
            return
        elif sys.argv[1] == '--template':
            # 生成Excel模板
            path = generate_excel_template()
            if path:
                print_rich(f"[green]✅ Excel模板已生成: {path}[/green]")
            return
    
    # 交互式菜单
    while True:
        show_menu()
        choice = input_with_default("请选择", "1")
        
        if choice == '1':
            # 交互式配置
            info = gather_info_interactive()
            preview_config(info)
            if Confirm.ask("确认执行适配？"):
                execute_adaptation(info)
                show_completion(info)
                break
            
        elif choice == '2':
            # Excel导入
            if not HAS_OPENPYXL:
                print_rich("[red]错误：需要安装 openpyxl 库[/red]")
                print_rich("[yellow]请运行：pip install openpyxl[/yellow]")
                continue
            
            # 询问是否生成模板
            if Confirm.ask("是否先下载Excel模板？"):
                path = generate_excel_template()
                if path:
                    print_rich(f"[green]✅ 模板已生成: {path}[/green]")
                    print_rich("[yellow]请填写模板后，输入文件路径导入[/yellow]")
            
            excel_path = input_with_default("请输入Excel文件路径", "")
            if excel_path and Path(excel_path).exists():
                info = import_from_excel(excel_path)
                if info:
                    preview_config(info)
                    if Confirm.ask("确认执行适配？"):
                        execute_adaptation(info)
                        show_completion(info)
                        break
            else:
                print_rich("[red]文件不存在，请检查路径[/red]")
                
        elif choice == '3':
            # JSON导入
            json_path = input_with_default("请输入JSON配置文件路径", "")
            if json_path and Path(json_path).exists():
                info = import_from_json(json_path)
                preview_config(info)
                if Confirm.ask("确认执行适配？"):
                    execute_adaptation(info)
                    show_completion(info)
                    break
            else:
                print_rich("[red]文件不存在，请检查路径[/red]")
                
        elif choice == '4':
            print_rich("[yellow]已退出[/yellow]")
            break
        else:
            print_rich("[red]无效选择，请重试[/red]")

if __name__ == '__main__':
    main()
